#!/usr/bin/env python3
"""
build_dashboard.py — Comfrut Dashboard Builder
Lee el Excel de data/, procesa los datos y los embebe en el template HTML.
Genera docs/index.html listo para GitHub Pages.

Hojas requeridas en el Excel:
  - "Asistencia + Produccion" → PROD + TD
  - "Tiempos Perdidos"        → PD
  - "Programa Envasado"       → PROG (formato wide: fila0=turnos, fila1=fechas, fila2+=cods)
"""

import json
import math
import os
import sys
import glob
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DATA_DIR      = os.path.join(BASE_DIR, "data")
TEMPLATE_DIR  = os.path.join(BASE_DIR, "templates")
OUTPUT_DIR    = os.path.join(BASE_DIR, "docs")
TEMPLATE_FILE = os.path.join(TEMPLATE_DIR, "template.html")
OUTPUT_FILE   = os.path.join(OUTPUT_DIR, "index.html")

# ─── Helpers ─────────────────────────────────────────────────────────────────
def fmt_fecha(cell_value):
    """Excel cell → 'YYYY-MM-DD' string."""
    if cell_value is None:
        return ""
    if isinstance(cell_value, (datetime, date)):
        return cell_value.strftime("%Y-%m-%d")
    if isinstance(cell_value, (int, float)):
        # Excel serial date
        try:
            d = datetime(1899, 12, 30) + __import__('datetime').timedelta(days=int(cell_value))
            return d.strftime("%Y-%m-%d")
        except Exception:
            return ""
    s = str(cell_value).strip()
    return s[:10] if len(s) >= 10 else ""

def safe_float(v, default=0.0):
    try:
        return float(v) if v not in (None, "") else default
    except (ValueError, TypeError):
        return default

def safe_int(v, default=0):
    try:
        return int(float(v)) if v not in (None, "") else default
    except (ValueError, TypeError):
        return default

def safe_str(v, default=""):
    return str(v).strip() if v not in (None, "") else default

def iso_week(fecha_str):
    """'YYYY-MM-DD' → ISO week number (JS compatible formula)."""
    try:
        d = datetime.strptime(fecha_str, "%Y-%m-%d")
        jan1 = datetime(d.year, 1, 1)
        diff_days = (d - jan1).days
        jan1_js_day = jan1.isoweekday() % 7  # JS: 0=sun,1=mon,...
        return math.ceil((diff_days + jan1_js_day + 1) / 7)
    except Exception:
        return 0

# ─── Find Excel ───────────────────────────────────────────────────────────────
def find_excel():
    patterns = ["*.xlsx", "*.xlsm", "*.xls"]
    for pat in patterns:
        files = glob.glob(os.path.join(DATA_DIR, pat))
        if files:
            # Use most recently modified
            files.sort(key=os.path.getmtime, reverse=True)
            return files[0]
    return None

# ─── Sheet parsers ────────────────────────────────────────────────────────────
def parse_asistencia_prod(ws):
    """
    Hoja 'Asistencia + Produccion'.
    Retorna PROD (lista de registros) y TD (lista de tiempos teóricos).
    """
    headers = {}
    PROD = []
    TD   = []
    
    for row in ws.iter_rows():
        if not headers:
            # Detect header row
            row_vals = [str(c.value).strip() if c.value else "" for c in row]
            if "Inic.tratamiento" in row_vals or "Pto. Trabajo" in row_vals:
                headers = {v: i for i, v in enumerate(row_vals) if v}
            continue

        def g(col, default=""):
            idx = headers.get(col)
            if idx is None:
                return default
            v = row[idx].value
            return v if v is not None else default

        fecha = fmt_fecha(g("Inic.tratamiento"))
        if not fecha or len(fecha) < 10:
            continue

        try:
            dt = datetime.strptime(fecha, "%Y-%m-%d")
        except ValueError:
            continue

        linea = safe_str(g("Pto. Trabajo"))
        area  = "ZENV" if linea.startswith("S_ENV") else "ZCIQ"
        sem   = safe_int(g("Semana"))

        prod_row = {
            "fecha":        fecha,
            "año":          dt.year,
            "mes":          dt.month,
            "dia":          dt.day,
            "semana":       sem,
            "turno":        safe_int(g("Turno"), 1),
            "linea":        linea,
            "area":         area,
            "especie":      safe_str(g("Especie")),
            "sku":          safe_str(g("Desc.Material")),
            "cod":          safe_str(g("Material")),
            "ton_real":     safe_float(g("Ton.Real")),
            "cajas":        safe_float(g("Cajas Produc.")),
            "bpm_total":    safe_float(g("BPM Total")),
            "bpm_std":      safe_float(g("BPM Estandar")),
            "kg_ingresados":safe_float(g("Kilos Ingresados")),
            "iqf_aprobado": safe_float(g("IQF Aprobado")),
            "kg_aprobadas": safe_float(g("Kilos Aprobadas")),
            "kg_puro":      safe_float(g("Kilos Pure")),
            "kg_jugo":      safe_float(g("Kilos Jugo")),
            "kg_crumble":   safe_float(g("Kilos Crumble")),
            "kg_h_pers":    safe_float(g("Produc.(Kg/H/Personas)")),
            "personas":     safe_float(g("Cant.Personas")),
            "teo_cajas":    safe_float(g("Teorico Cajas")),
            "con_cajas":    safe_float(g("Consumo Cajas")),
            "teo_bolsas":   safe_float(g("Teorico Bolsas")),
            "con_bolsas":   safe_float(g("Consumo Bolsas")),
            "teo_min":      safe_float(g("T.Minutos")),
            "efec_min":     safe_float(g("Tiempo Efec.Min.")),
        }
        PROD.append(prod_row)

        td_row = {
            "fecha":    fecha,
            "año":      dt.year,
            "mes":      dt.month,
            "dia":      dt.day,
            "semana":   sem,
            "linea":    linea,
            "minutos":  safe_float(g("T.Minutos")),
            "efec_min": safe_float(g("Tiempo Efec.Min.")),
            "plan_min": safe_float(g("Paros Plan Min.")),
            "nopl_min": safe_float(g("Paros No Plan Min.")),
        }
        TD.append(td_row)

    return PROD, TD


def parse_tiempos_perdidos(ws):
    """Hoja 'Tiempos Perdidos' → PD."""
    headers = {}
    PD = []

    for row in ws.iter_rows():
        if not headers:
            row_vals = [str(c.value).strip() if c.value else "" for c in row]
            if "Fecha" in row_vals and "T.Minutos" in row_vals:
                headers = {v: i for i, v in enumerate(row_vals) if v}
            continue

        def g(col, default=""):
            idx = headers.get(col)
            if idx is None:
                return default
            v = row[idx].value
            return v if v is not None else default

        fecha = fmt_fecha(g("Fecha"))
        if not fecha or len(fecha) < 10:
            continue

        try:
            dt = datetime.strptime(fecha, "%Y-%m-%d")
        except ValueError:
            continue

        PD.append({
            "fecha":     fecha,
            "año":       dt.year,
            "mes":       dt.month,
            "dia":       dt.day,
            "semana":    safe_int(g("Semana")),
            "turno":     safe_int(g("Turno"), 1),
            "linea":     safe_str(g("Pto. Trabajo")),
            "area":      safe_str(g("Clase de Orden"), "ZCIQ"),
            "tipo_paro": safe_str(g("Tipo de Paro")),
            "categoria": safe_str(g("Desc.Clasifi. del Paro"), "Producción"),
            "falla":     safe_str(g("Desc.Falla"), "SIN DESCRIPCIÓN"),
            "obs":       safe_str(g("Observaciones")),
            "minutos":   safe_float(g("T.Minutos")),
        })

    return PD


def parse_programa_envasado(ws):
    """
    Hoja 'Programa Envasado' (formato wide):
      fila 0: [None, None, "Turno", 3, 1, 2, 3, 1, 2, ...]
      fila 1: [None, None, None,  fecha, fecha, fecha, ...]
      fila 2+: [cod, sku, peso, val, val, val, ...]
    Retorna lista plana de {fecha, cod, turno, cajas_prog}.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    turno_row = rows[0]   # turnos en col 3+
    fecha_row = rows[1]   # fechas en col 3+
    PROG = []

    for row in rows[2:]:
        cod = row[0]
        if cod is None or str(cod).strip() == "":
            continue
        cod = str(cod).strip()

        for ci in range(3, len(fecha_row)):
            val = row[ci] if ci < len(row) else None
            if val is None or val == 0 or val == "":
                continue
            try:
                num_val = float(val)
            except (ValueError, TypeError):
                continue
            if num_val <= 0:
                continue

            raw_fecha = fecha_row[ci]
            raw_turno = turno_row[ci]
            if raw_fecha is None or raw_turno is None:
                continue

            fecha_str = fmt_fecha(raw_fecha)
            if not fecha_str or len(fecha_str) < 10:
                continue

            try:
                turno = int(float(raw_turno))
            except (ValueError, TypeError):
                continue

            PROG.append({
                "fecha":      fecha_str,
                "cod":        cod,
                "turno":      turno,
                "cajas_prog": num_val,
            })

    return PROG


def build_indices(PD, PROD):
    """Construye PIDX y IDX (cascade filter indices) compatibles con el JS."""
    ZCIQ_L = {"S_L01", "S_L03", "S_L04", "S_L05", "S_LMANCU", "S_MANUAL"}
    ZENV_L = {"S_ENV1"}

    def make_idx():
        return {"MS": {}, "MD": {}, "SD": {}}

    def add_to(obj, key, val):
        if key not in obj:
            obj[key] = []
        if val not in obj[key]:
            obj[key].append(val)

    # TP indices (from PD)
    tp = {
        "MS": {}, "MD": {}, "SD": {},
        "MS_ZCIQ": {}, "MD_ZCIQ": {}, "SD_ZCIQ": {},
        "MS_ZENV": {}, "MD_ZENV": {}, "SD_ZENV": {},
    }
    for r in PD:
        mk = str(r["mes"]); sk = str(r["semana"])
        add_to(tp["MS"], mk, r["semana"]); add_to(tp["MD"], mk, r["dia"]); add_to(tp["SD"], sk, r["dia"])
        if r["linea"] in ZCIQ_L:
            add_to(tp["MS_ZCIQ"], mk, r["semana"]); add_to(tp["MD_ZCIQ"], mk, r["dia"]); add_to(tp["SD_ZCIQ"], sk, r["dia"])
        if r["linea"] in ZENV_L:
            add_to(tp["MS_ZENV"], mk, r["semana"]); add_to(tp["MD_ZENV"], mk, r["dia"]); add_to(tp["SD_ZENV"], sk, r["dia"])

    for obj in tp.values():
        for k in obj:
            obj[k] = sorted(set(obj[k]))

    # PROD indices (from PROD, same structure)
    pr = {
        "MS": {}, "MD": {}, "SD": {},
        "MS_ZCIQ": {}, "MD_ZCIQ": {}, "SD_ZCIQ": {},
        "MS_ZENV": {}, "MD_ZENV": {}, "SD_ZENV": {},
    }
    for r in PROD:
        mk = str(r["mes"]); sk = str(r["semana"])
        add_to(pr["MS"], mk, r["semana"]); add_to(pr["MD"], mk, r["dia"]); add_to(pr["SD"], sk, r["dia"])
        if r["linea"] in ZCIQ_L:
            add_to(pr["MS_ZCIQ"], mk, r["semana"]); add_to(pr["MD_ZCIQ"], mk, r["dia"]); add_to(pr["SD_ZCIQ"], sk, r["dia"])
        if r["linea"] in ZENV_L:
            add_to(pr["MS_ZENV"], mk, r["semana"]); add_to(pr["MD_ZENV"], mk, r["dia"]); add_to(pr["SD_ZENV"], sk, r["dia"])

    for obj in pr.values():
        for k in obj:
            obj[k] = sorted(set(obj[k]))

    return pr, tp  # PIDX, IDX


def embed_data(template_html, PROD, PIDX, PROG, IDX):
    """Reemplaza los null en el template con los datos reales."""
    def replace_raw(html, name, data):
        marker_start = f"const {name} = (function(){{ try{{ return "
        marker_end   = "; }catch(e){ return null; } })();"
        idx_s = html.find(marker_start)
        if idx_s == -1:
            print(f"  WARN: {name} not found in template")
            return html
        data_start = idx_s + len(marker_start)
        idx_e = html.find(marker_end, data_start)
        if idx_e == -1:
            print(f"  WARN: {name} end marker not found")
            return html
        json_str = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
        return html[:data_start] + json_str + html[idx_e:]

    html = template_html
    html = replace_raw(html, "_PROD_RAW", PROD)
    html = replace_raw(html, "_PIDX_RAW", PIDX)
    html = replace_raw(html, "_PROG_RAW", PROG)
    html = replace_raw(html, "_IDX_RAW",  IDX)
    return html


# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  Comfrut Dashboard Builder")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 1. Find Excel
    excel_path = find_excel()
    if not excel_path:
        print(f"ERROR: No se encontró ningún Excel en {DATA_DIR}/")
        print("       Sube el archivo .xlsx a la carpeta data/")
        sys.exit(1)
    print(f"\n📊 Excel: {os.path.basename(excel_path)}")

    # 2. Load workbook
    print("   Cargando workbook...")
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"ERROR cargando Excel: {e}")
        sys.exit(1)

    print(f"   Hojas disponibles: {wb.sheetnames}")

    # 3. Find sheets
    def find_sheet(keywords):
        for name in wb.sheetnames:
            nl = name.lower()
            if all(k in nl for k in keywords):
                return name
        # Partial match
        for name in wb.sheetnames:
            nl = name.lower()
            if any(k in nl for k in keywords):
                return name
        return None

    sh_ap   = find_sheet(["asistencia"]) or find_sheet(["produccion"])
    sh_tp   = find_sheet(["tiempos"])    or find_sheet(["perdid"])
    sh_prog = find_sheet(["programa"])   or find_sheet(["envas"])

    print(f"\n   Hoja producción   : {sh_ap}")
    print(f"   Hoja tiempos      : {sh_tp}")
    print(f"   Hoja programa     : {sh_prog}")

    # 4. Parse sheets
    PROD, TD = [], []
    if sh_ap:
        print("\n⚙️  Procesando Asistencia + Producción...")
        PROD, TD = parse_asistencia_prod(wb[sh_ap])
        print(f"   PROD: {len(PROD)} registros")
        print(f"   TD:   {len(TD)} registros")
    else:
        print("WARN: Hoja 'Asistencia + Produccion' no encontrada")

    PD = []
    if sh_tp:
        print("\n⚙️  Procesando Tiempos Perdidos...")
        PD = parse_tiempos_perdidos(wb[sh_tp])
        print(f"   PD: {len(PD)} registros")
    else:
        print("WARN: Hoja 'Tiempos Perdidos' no encontrada")

    PROG = []
    if sh_prog:
        print("\n⚙️  Procesando Programa Envasado...")
        PROG = parse_programa_envasado(wb[sh_prog])
        print(f"   PROG: {len(PROG)} registros")
    else:
        print("WARN: Hoja 'Programa Envasado' no encontrada")

    # 5. Build indices
    print("\n⚙️  Construyendo índices...")
    PIDX, IDX = build_indices(PD, PROD)
    print(f"   PIDX semanas ZENV: {list(PIDX.get('MS_ZENV', {}).keys())[:5]}...")
    print(f"   IDX  semanas TP  : {list(IDX.get('MS', {}).keys())[:5]}...")

    # 6. Load template
    if not os.path.exists(TEMPLATE_FILE):
        print(f"\nERROR: Template no encontrado en {TEMPLATE_FILE}")
        print("       Asegúrate de subir template.html a la carpeta templates/")
        sys.exit(1)

    print(f"\n📄 Template: {TEMPLATE_FILE}")
    with open(TEMPLATE_FILE, "r", encoding="utf-8") as f:
        template_html = f.read()

    # 7. Embed data
    print("\n⚙️  Embebiendo datos en template...")
    output_html = embed_data(template_html, PROD, PIDX, PROG, IDX)
    
    # Verify embedding worked
    if '"fecha"' not in output_html and len(PROD) > 0:
        print("ERROR: Embedding falló — los datos no aparecen en el HTML")
        sys.exit(1)

    # 8. Write output
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(output_html)

    size_kb = os.path.getsize(OUTPUT_FILE) // 1024
    print(f"\n✅ Generado: {OUTPUT_FILE}")
    print(f"   Tamaño: {size_kb} KB")
    print(f"   PROD: {len(PROD)} · PD: {len(PD)} · PROG: {len(PROG)}")
    print(f"\n🚀 GitHub Pages se actualizará en ~1-2 minutos.")
    print("=" * 60)


if __name__ == "__main__":
    main()
